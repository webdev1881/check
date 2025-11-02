import asyncio
import aiohttp
import ssl
import pandas as pd
import random
from typing import List, Dict, Tuple
from dataclasses import dataclass
import logging
from pathlib import Path


# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('discount_checker.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class Config:
    BASE_URL = "https://89.105.216.114"
    USERNAME = "Yulia"
    PASSWORD = "SY1804$@"
    
    BATCH_SIZE = 100
    USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36'
    
    EXCEL_FILE = "data.xlsx"  # –ò–º—è Excel —Ñ–∞–π–ª–∞ –≤ –∫–æ—Ä–Ω–µ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è


@dataclass
class RuleSet:
    """–ù–∞–±–æ—Ä –ø—Ä–∞–≤–∏–ª –¥–ª—è –æ–¥–Ω–æ–π —Å—Ç—Ä–æ–∫–∏"""
    article: str  # –ê—Ä—Ç–∏–∫—É–ª –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ C
    price: float  # –¶–µ–Ω–∞ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ I
    
    # –£—Ä–æ–≤–Ω–∏
    level_0: float  # –°–ª—É—á–∞–π–Ω–æ–µ —á–∏—Å–ª–æ –æ—Ç 0 –¥–æ K (–Ω–µ –≤–∫–ª—é—á–∞—è)
    level_1: float  # –°–ª—É—á–∞–π–Ω–æ–µ —á–∏—Å–ª–æ –æ—Ç K (–≤–∫–ª—é—á–∞—è) –¥–æ P (–Ω–µ –≤–∫–ª—é—á–∞—è)
    level_2: float  # –°–ª—É—á–∞–π–Ω–æ–µ —á–∏—Å–ª–æ –≤—ã—à–µ –Ω–∞ 150-300% –æ—Ç P (–≤–∫–ª—é—á–∞—è)
    
    # –ü—Ä–∞–≤–∏–ª–∞
    rule_0: float  # level_0 * price
    rule_1: float  # (price - L) * level_1
    rule_1_1: float  # Q * K
    rule_2: float  # (price - P) * level_2
    rule_2_1: float  # Q * P
    
    # –ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ –¥–ª—è —Å–ø—Ä–∞–≤–∫–∏
    k_value: float
    l_value: float
    p_value: float
    q_value: float


class ExcelParser:
    """–ü–∞—Ä—Å–µ—Ä Excel —Ñ–∞–π–ª–∞"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        
    def parse(self) -> List[RuleSet]:
        """–ü–∞—Ä—Å–∏—Ç Excel –∏ —Å–æ–∑–¥–∞–µ—Ç –Ω–∞–±–æ—Ä—ã –ø—Ä–∞–≤–∏–ª –¥–ª—è –∫–∞–∂–¥–æ–π —Å—Ç—Ä–æ–∫–∏"""
        try:
            # –ß–∏—Ç–∞–µ–º Excel —Ñ–∞–π–ª
            df = pd.read_excel(self.file_path)
            logger.info(f"Excel —Ñ–∞–π–ª –∑–∞–≥—Ä—É–∂–µ–Ω: {len(df)} —Å—Ç—Ä–æ–∫")
            
            rule_sets = []
            
            for index, row in df.iterrows():
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Å—Ç–æ–ª–±–µ—Ü C –Ω–µ –ø—É—Å—Ç–æ–π
                if pd.isna(row.iloc[2]) or str(row.iloc[2]).strip() == '':
                    continue
                
                try:
                    # –ò–∑–≤–ª–µ–∫–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ —Å—Ç–æ–ª–±—Ü–æ–≤
                    article = str(row.iloc[2]).strip()  # –°—Ç–æ–ª–±–µ—Ü C
                    price = float(row.iloc[8])  # –°—Ç–æ–ª–±–µ—Ü I
                    k_value = float(row.iloc[10])  # –°—Ç–æ–ª–±–µ—Ü K
                    l_value = float(row.iloc[11])  # –°—Ç–æ–ª–±–µ—Ü L
                    p_value = float(row.iloc[15])  # –°—Ç–æ–ª–±–µ—Ü P
                    q_value = float(row.iloc[16])  # –°—Ç–æ–ª–±–µ—Ü Q
                    
                    # –°–æ–∑–¥–∞–µ–º —É—Ä–æ–≤–Ω–∏ (–ø—Ä–µ–¥–≤–∞—Ä–∏—Ç–µ–ª—å–Ω—ã–µ —Ä–∞—Å—á–µ—Ç—ã)
                    level_0 = round(random.uniform(0, k_value), 2)
                    level_1 = round(random.uniform(k_value, p_value), 2)
                    level_2 = round(p_value * random.uniform(1.5, 3.0), 2)
                    
                    # –ü—Ä–∏–º–µ–Ω—è–µ–º –ø—Ä–∞–≤–∏–ª–∞
                    rule_0 = round(level_0 * price, 2)
                    rule_1 = round((price - l_value) * level_1, 2)
                    rule_1_1 = round((price - l_value) * k_value, 2)
                    rule_2 = round((price - q_value) * level_2, 2)
                    rule_2_1 = round((price - q_value) * p_value, 2)
                    
                    rule_set = RuleSet(
                        article=article,
                        price=price,
                        level_0=level_0,
                        level_1=level_1,
                        level_2=level_2,
                        rule_0=rule_0,
                        rule_1=rule_1,
                        rule_1_1=rule_1_1,
                        rule_2=rule_2,
                        rule_2_1=rule_2_1,
                        k_value=k_value,
                        l_value=l_value,
                        p_value=p_value,
                        q_value=q_value
                    )
                    
                    rule_sets.append(rule_set)
                    logger.debug(f"–°–æ–∑–¥–∞–Ω –Ω–∞–±–æ—Ä –ø—Ä–∞–≤–∏–ª –¥–ª—è –∞—Ä—Ç–∏–∫—É–ª–∞ {article}")
                    
                except (ValueError, IndexError) as e:
                    logger.warning(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å—Ç—Ä–æ–∫–∏ {index + 1}: {e}")
                    continue
            
            logger.info(f"–°–æ–∑–¥–∞–Ω–æ {len(rule_sets)} –Ω–∞–±–æ—Ä–æ–≤ –ø—Ä–∞–≤–∏–ª")
            return rule_sets
            
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è Excel —Ñ–∞–π–ª–∞: {e}")
            raise


class DiscountRulesAPI:
    """API –∫–ª–∏–µ–Ω—Ç –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å —Å–∏—Å—Ç–µ–º–æ–π —Å–∫–∏–¥–æ–∫"""
    
    def __init__(self, config: Config):
        self.config = config
        self.session = None
        self.cookies = None
        
    async def __aenter__(self):
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        self.session = aiohttp.ClientSession(connector=connector)
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            await self.session.close()
            await asyncio.sleep(0.25)  # –î–∞–µ–º –≤—Ä–µ–º—è –Ω–∞ –∑–∞–∫—Ä—ã—Ç–∏–µ
    
    async def login(self) -> bool:
        """–ê–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—è –≤ —Å–∏—Å—Ç–µ–º–µ"""
        url = f"{self.config.BASE_URL}/api/login"
        payload = {
            "username": self.config.USERNAME,
            "password": self.config.PASSWORD
        }
        
        try:
            async with self.session.post(url, json=payload) as response:
                if response.status == 200:
                    self.cookies = response.cookies
                    logger.info("–ê–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—è —É—Å–ø–µ—à–Ω–∞")
                    return True
                else:
                    text = await response.text()
                    logger.error(f"–û—à–∏–±–∫–∞ –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏: {response.status} - {text}")
                    return False
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏: {e}")
            return False
    
    async def get_discount_rules_page(self, offset: int = 0) -> Tuple[List[Dict], int]:
        """–ü–æ–ª—É—á–∞–µ—Ç –æ–¥–Ω—É —Å—Ç—Ä–∞–Ω–∏—Ü—É –ø—Ä–∞–≤–∏–ª —Å–∫–∏–¥–æ–∫"""
        url = f"{self.config.BASE_URL}/discountRule/list"
        
        payload = {
            "count": self.config.BATCH_SIZE,
            "filter": {},
            "offset": offset,
            "period": {},
            "sort": {
                "fields": [
                    {
                        "field": "name",
                        "asc": True
                    }
                ]
            }
        }
        
        headers = {
            'accept': '*/*',
            'content-type': 'application/json',
            'origin': self.config.BASE_URL,
            'referer': f"{self.config.BASE_URL}/",
            'user-agent': self.config.USER_AGENT
        }
        
        try:
            async with self.session.post(url, json=payload, headers=headers, cookies=self.cookies) as response:
                if response.status == 200:
                    data = await response.json()
                    return data.get('data', []), data.get('count', 0)
                else:
                    text = await response.text()
                    logger.error(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö: {response.status} - {text}")
                    return [], 0
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø—Ä–æ—Å–µ –¥–∞–Ω–Ω—ã—Ö: {e}")
            return [], 0
    
    async def get_all_discount_rules(self) -> List[Dict]:
        """–ü–æ–ª—É—á–∞–µ—Ç –≤—Å–µ –ø—Ä–∞–≤–∏–ª–∞ —Å–∫–∏–¥–æ–∫ —Å —É—á–µ—Ç–æ–º –ø–∞–≥–∏–Ω–∞—Ü–∏–∏"""
        all_rules = []
        offset = 0
        
        while True:
            rules, total_count = await self.get_discount_rules_page(offset)
            
            if not rules:
                break
            
            all_rules.extend(rules)
            logger.info(f"–ó–∞–≥—Ä—É–∂–µ–Ω–æ {len(all_rules)} –∏–∑ {total_count} –ø—Ä–∞–≤–∏–ª")
            
            if len(all_rules) >= total_count:
                break
            
            offset += self.config.BATCH_SIZE
        
        logger.info(f"–í—Å–µ–≥–æ –∑–∞–≥—Ä—É–∂–µ–Ω–æ {len(all_rules)} –ø—Ä–∞–≤–∏–ª")
        return all_rules
    
    async def find_rules_by_articles(self, articles: List[str]) -> Dict[str, List[Dict]]:
        """–ù–∞—Ö–æ–¥–∏—Ç –ø—Ä–∞–≤–∏–ª–∞ –¥–ª—è —Å–ø–∏—Å–∫–∞ –∞—Ä—Ç–∏–∫—É–ª–æ–≤"""
        all_rules = await self.get_all_discount_rules()
        
        # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º –ø—Ä–∞–≤–∏–ª–∞ –ø–æ –∞—Ä—Ç–∏–∫—É–ª–∞–º
        rules_by_article = {article: [] for article in articles}
        
        for rule in all_rules:
            name = rule.get('name', '')
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É "–ê—Ö—Ç–∏—Ä–∫–∞_{Article}"
            if name.startswith('–ê—Ö—Ç–∏—Ä–∫–∞_'):
                article = name.split('–ê—Ö—Ç–∏—Ä–∫–∞_', 1)[1]
                
                if article in rules_by_article:
                    rules_by_article[article].append(rule)
        
        # –õ–æ–≥–∏—Ä—É–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        for article, rules in rules_by_article.items():
            if rules:
                logger.info(f"–î–ª—è –∞—Ä—Ç–∏–∫—É–ª–∞ {article} –Ω–∞–π–¥–µ–Ω–æ {len(rules)} –ø—Ä–∞–≤–∏–ª")
            else:
                logger.warning(f"–î–ª—è –∞—Ä—Ç–∏–∫—É–ª–∞ {article} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø—Ä–∞–≤–∏–ª")
        
        return rules_by_article
    
    async def test_discount_rule(self, article: str, quantity: float, price: float, terminal_id: int = 1541) -> Dict:
        """–¢–µ—Å—Ç–∏—Ä—É–µ—Ç –ø—Ä–∞–≤–∏–ª–æ —Å–∫–∏–¥–∫–∏"""
        url = f"{self.config.BASE_URL}/discountRuleTester/process"
        
        payload = {
            "items": [
                {
                    "extSku": {
                        "id": article
                    },
                    "quantity": quantity,
                    "price": str(price),
                    "discount": 0,
                    "coupons": [],
                    "paidByPoints": None,
                    "appliedDiscountAmount": None,
                    "isFullTank": False,
                    "amount": round(quantity * price, 2)
                }
            ],
            "promoCodes": "",
            "cardCode": None,
            "clientId": None,
            "payFormType": 0,
            "terminalId": terminal_id,
            "date": "2025-11-01T16:46:39.609Z"
        }
        
        headers = {
            'accept': '*/*',
            'content-type': 'application/json',
            'origin': self.config.BASE_URL,
            'referer': f"{self.config.BASE_URL}/",
            'user-agent': self.config.USER_AGENT
        }
        
        try:
            async with self.session.post(url, json=payload, headers=headers, cookies=self.cookies) as response:
                if response.status == 200:
                    data = await response.json()
                    return {
                        'success': True,
                        'data': data,
                        'total_discount': data.get('data', {}).get('totalDiscountAmount', 0)
                    }
                else:
                    text = await response.text()
                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞ –æ—à–∏–±–∫—É –ë–î - –∞—Ä—Ç–∏–∫—É–ª –Ω–µ –Ω–∞–π–¥–µ–Ω
                    if 'is not present in table' in text or 'ext_sku_group_id' in text:
                        return {
                            'success': False,
                            'error': '–ê—Ä—Ç–∏–∫—É–ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–∏—Å—Ç–µ–º–µ',
                            'total_discount': 0
                        }
                    logger.error(f"–û—à–∏–±–∫–∞ —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –ø—Ä–∞–≤–∏–ª–∞: {response.status} - {text}")
                    return {
                        'success': False,
                        'error': text[:200],  # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –¥–ª–∏–Ω—É
                        'total_discount': 0
                    }
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ –ø—Ä–∞–≤–∏–ª–∞: {e}")
            return {
                'success': False,
                'error': str(e)[:200],
                'total_discount': 0
            }


@dataclass
class ValidationCheck:
    """–†–µ–∑—É–ª—å—Ç–∞—Ç –ø—Ä–æ–≤–µ—Ä–∫–∏ –æ–¥–Ω–æ–≥–æ –ø—Ä–∞–≤–∏–ª–∞"""
    rule_name: str
    quantity: float
    price_without_discount: float
    price_with_discount: float
    expected_discount: float
    actual_discount: float
    difference: float
    status: str  # 'OK', 'FAIL', 'ERROR'
    error: str = None


class RulesValidator:
    """–í–∞–ª–∏–¥–∞—Ç–æ—Ä –ø—Ä–∞–≤–∏–ª"""
    
    def __init__(self, api: DiscountRulesAPI, terminal_id: int = 1541):
        self.api = api
        self.terminal_id = terminal_id
        self.results = []
    
    async def validate(self, rule_set: RuleSet, api_rules: List[Dict]) -> Dict:
        """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –ø—Ä–∞–≤–∏–ª–∞ –∏–∑ API –ø—Ä–æ—Ç–∏–≤ —Ä–∞—Å—á–µ—Ç–Ω—ã—Ö –ø—Ä–∞–≤–∏–ª"""
        validation_result = {
            'article': rule_set.article,
            'price': rule_set.price,
            'api_rules_count': len(api_rules),
            'checks': []
        }
        
        if not api_rules:
            validation_result['status'] = 'NO_API_RULES'
            validation_result['message'] = '–ü—Ä–∞–≤–∏–ª–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã –≤ API'
            return validation_result
        
        logger.info(f"\n{'='*80}")
        logger.info(f"üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ –∞—Ä—Ç–∏–∫—É–ª–∞: {rule_set.article}")
        logger.info(f"{'='*80}")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥–æ–µ –∏–∑ 5 –ø—Ä–∞–≤–∏–ª
        rules_to_check = [
            ('–ü—Ä–∞–≤–∏–ª–æ 0', rule_set.level_0, rule_set.rule_0),
            ('–ü—Ä–∞–≤–∏–ª–æ 1', rule_set.level_1, rule_set.rule_1),
            ('–ü—Ä–∞–≤–∏–ª–æ 1-1', rule_set.k_value, rule_set.rule_1_1),
            ('–ü—Ä–∞–≤–∏–ª–æ 2', rule_set.level_2, rule_set.rule_2),
            ('–ü—Ä–∞–≤–∏–ª–æ 2-1', rule_set.p_value, rule_set.rule_2_1),
        ]
        
        for rule_name, quantity, price_with_discount in rules_to_check:
            # –¶–µ–Ω–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏
            price_without_discount = round(quantity * rule_set.price, 2)
            
            # –û–∂–∏–¥–∞–µ–º–∞—è —Å–∫–∏–¥–∫–∞ = —Ü–µ–Ω–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏ - —Ü–µ–Ω–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π
            expected_discount = round(price_without_discount - price_with_discount, 2)
            
            logger.info(f"\nüìã {rule_name}:")
            logger.info(f"   –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ: {quantity}")
            logger.info(f"   –¶–µ–Ω–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏: {price_without_discount}")
            logger.info(f"   –¶–µ–Ω–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π: {price_with_discount}")
            logger.info(f"   –û–∂–∏–¥–∞–µ–º–∞—è —Å–∫–∏–¥–∫–∞: {expected_discount}")
            
            # –¢–µ—Å—Ç–∏—Ä—É–µ–º —á–µ—Ä–µ–∑ API
            result = await self.api.test_discount_rule(
                article=rule_set.article,
                quantity=quantity,
                price=rule_set.price,
                terminal_id=self.terminal_id
            )
            
            if result['success']:
                actual_discount = result['total_discount']
                difference = abs(expected_discount - actual_discount)
                
                # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Å—Ç–∞—Ç—É—Å (–¥–æ–ø—É—Å–∫ 0.01)
                status = 'OK' if difference <= 0.01 else 'FAIL'
                
                check = ValidationCheck(
                    rule_name=rule_name,
                    quantity=quantity,
                    price_without_discount=price_without_discount,
                    price_with_discount=price_with_discount,
                    expected_discount=expected_discount,
                    actual_discount=actual_discount,
                    difference=difference,
                    status=status
                )
                
                # –ö—Ä–∞—Å–∏–≤—ã–π –≤—ã–≤–æ–¥
                if status == 'OK':
                    logger.info(f"   ‚úÖ API —Å–∫–∏–¥–∫–∞: {actual_discount} - –°–û–í–ü–ê–î–ê–ï–¢")
                else:
                    logger.warning(f"   ‚ùå API —Å–∫–∏–¥–∫–∞: {actual_discount} - –†–ê–°–•–û–ñ–î–ï–ù–ò–ï {difference}")
                
            else:
                check = ValidationCheck(
                    rule_name=rule_name,
                    quantity=quantity,
                    price_without_discount=price_without_discount,
                    price_with_discount=price_with_discount,
                    expected_discount=expected_discount,
                    actual_discount=0,
                    difference=expected_discount,
                    status='ERROR',
                    error=result.get('error', 'Unknown error')
                )
                logger.error(f"   ‚ùå –û—à–∏–±–∫–∞ API: {check.error}")
            
            validation_result['checks'].append(check)
        
        # –ü–æ–¥—Å—á–µ—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏
        ok_count = sum(1 for c in validation_result['checks'] if c.status == 'OK')
        fail_count = sum(1 for c in validation_result['checks'] if c.status == 'FAIL')
        error_count = sum(1 for c in validation_result['checks'] if c.status == 'ERROR')
        
        validation_result['status'] = 'COMPLETED'
        validation_result['ok_count'] = ok_count
        validation_result['fail_count'] = fail_count
        validation_result['error_count'] = error_count
        validation_result['message'] = f'–ü—Ä–æ–≤–µ—Ä–µ–Ω–æ 5 –ø—Ä–∞–≤–∏–ª: ‚úÖ {ok_count} | ‚ùå {fail_count} | ‚ö†Ô∏è {error_count}'
        
        logger.info(f"\nüìä –ò—Ç–æ–≥: {validation_result['message']}")
        
        return validation_result
    
    def export_to_excel(self, filename: str = "validation_results.xlsx"):
        """–≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –≤ Excel"""
        rows = []
        
        for result in self.results:
            article = result['article']
            price = result['price']
            status = result['status']
            
            if status == 'NO_API_RULES':
                rows.append({
                    '–ê—Ä—Ç–∏–∫—É–ª': article,
                    '–¶–µ–Ω–∞': price,
                    '–°—Ç–∞—Ç—É—Å': '–ù–µ—Ç –ø—Ä–∞–≤–∏–ª –≤ API',
                    '–ü—Ä–∞–≤–∏–ª–æ': '',
                    '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ': '',
                    '–°—É–º–º–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏': '',
                    '–°—É–º–º–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π': '',
                    '–û–∂–∏–¥–∞–µ–º–∞—è —Å–∫–∏–¥–∫–∞': '',
                    '–§–∞–∫—Ç–∏—á–µ—Å–∫–∞—è —Å–∫–∏–¥–∫–∞ (API)': '',
                    '–†–∞—Å—Ö–æ–∂–¥–µ–Ω–∏–µ': '',
                    '–†–µ–∑—É–ª—å—Ç–∞—Ç': '',
                    '–û—à–∏–±–∫–∞': ''
                })
            else:
                for check in result['checks']:
                    # –í–æ—Å—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –¥–µ—Ç–∞–ª–∏ –∏–∑ check
                    quantity = check.quantity if hasattr(check, 'quantity') else ''
                    price_without = check.price_without_discount if hasattr(check, 'price_without_discount') else ''
                    price_with = check.price_with_discount if hasattr(check, 'price_with_discount') else ''
                    
                    rows.append({
                        '–ê—Ä—Ç–∏–∫—É–ª': article,
                        '–¶–µ–Ω–∞': price,
                        '–°—Ç–∞—Ç—É—Å': status,
                        '–ü—Ä–∞–≤–∏–ª–æ': check.rule_name,
                        '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ': quantity,
                        '–°—É–º–º–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏': price_without,
                        '–°—É–º–º–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π': price_with,
                        '–û–∂–∏–¥–∞–µ–º–∞—è —Å–∫–∏–¥–∫–∞': check.expected_discount,
                        '–§–∞–∫—Ç–∏—á–µ—Å–∫–∞—è —Å–∫–∏–¥–∫–∞ (API)': check.actual_discount,
                        '–†–∞—Å—Ö–æ–∂–¥–µ–Ω–∏–µ': check.difference,
                        '–†–µ–∑—É–ª—å—Ç–∞—Ç': check.status,
                        '–û—à–∏–±–∫–∞': check.error if check.error else ''
                    })
        
        df = pd.DataFrame(rows)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='–†–µ–∑—É–ª—å—Ç–∞—Ç—ã')
            
            # –ü–æ–ª—É—á–∞–µ–º worksheet –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
            worksheet = writer.sheets['–†–µ–∑—É–ª—å—Ç–∞—Ç—ã']
            
            # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —à–∏—Ä–∏–Ω—É –∫–æ–ª–æ–Ω–æ–∫
            worksheet.column_dimensions['A'].width = 15  # –ê—Ä—Ç–∏–∫—É–ª
            worksheet.column_dimensions['B'].width = 10  # –¶–µ–Ω–∞
            worksheet.column_dimensions['C'].width = 20  # –°—Ç–∞—Ç—É—Å
            worksheet.column_dimensions['D'].width = 15  # –ü—Ä–∞–≤–∏–ª–æ
            worksheet.column_dimensions['E'].width = 12  # –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ
            worksheet.column_dimensions['F'].width = 18  # –°—É–º–º–∞ –±–µ–∑ —Å–∫–∏–¥–∫–∏
            worksheet.column_dimensions['G'].width = 18  # –°—É–º–º–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π
            worksheet.column_dimensions['H'].width = 18  # –û–∂–∏–¥–∞–µ–º–∞—è —Å–∫–∏–¥–∫–∞
            worksheet.column_dimensions['I'].width = 22  # –§–∞–∫—Ç–∏—á–µ—Å–∫–∞—è —Å–∫–∏–¥–∫–∞
            worksheet.column_dimensions['J'].width = 15  # –†–∞—Å—Ö–æ–∂–¥–µ–Ω–∏–µ
            worksheet.column_dimensions['K'].width = 12  # –†–µ–∑—É–ª—å—Ç–∞—Ç
            worksheet.column_dimensions['L'].width = 30  # –û—à–∏–±–∫–∞
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã (—Ü–≤–µ—Ç–∞–º–∏)
            for row in range(2, len(df) + 2):
                result_cell = worksheet.cell(row=row, column=11)  # –ö–æ–ª–æ–Ω–∫–∞ K (–†–µ–∑—É–ª—å—Ç–∞—Ç)
                
                if result_cell.value == 'OK':
                    result_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    result_cell.font = Font(color='006100')
                elif result_cell.value == 'FAIL':
                    result_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    result_cell.font = Font(color='9C0006')
                elif result_cell.value == 'ERROR':
                    result_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    result_cell.font = Font(color='9C6500')
        
        logger.info(f"\nüíæ –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤ {filename}")
        return filename


async def main():
    """–û—Å–Ω–æ–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –ø—Ä–æ–≥—Ä–∞–º–º—ã"""
    print("\n" + "="*80)
    print("üöÄ –ü–†–û–ì–†–ê–ú–ú–ê –ü–†–û–í–ï–†–ö–ò –ü–†–ê–í–ò–õ –°–ö–ò–î–û–ö")
    print("="*80)
    
    logger.info("="*80)
    logger.info("–ó–∞–ø—É—Å–∫ –ø—Ä–æ–≥—Ä–∞–º–º—ã –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø—Ä–∞–≤–∏–ª —Å–∫–∏–¥–æ–∫")
    logger.info("="*80)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ Excel —Ñ–∞–π–ª–∞
    excel_path = Path(Config.EXCEL_FILE)
    if not excel_path.exists():
        error_msg = f"‚ùå Excel —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {excel_path.absolute()}"
        print(error_msg)
        logger.error(error_msg)
        print("\nüí° –ü–æ–º–µ—Å—Ç–∏—Ç–µ —Ñ–∞–π–ª 'data.xlsx' –≤ –∫–æ—Ä–µ–Ω—å –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è")
        return
    
    # –ü–∞—Ä—Å–∏–º Excel –∏ —Å–æ–∑–¥–∞–µ–º –Ω–∞–±–æ—Ä—ã –ø—Ä–∞–≤–∏–ª
    print("\nüìÇ –®–∞–≥ 1: –ü–∞—Ä—Å–∏–Ω–≥ Excel —Ñ–∞–π–ª–∞...")
    logger.info("–®–∞–≥ 1: –ü–∞—Ä—Å–∏–Ω–≥ Excel —Ñ–∞–π–ª–∞")
    
    parser = ExcelParser(Config.EXCEL_FILE)
    rule_sets = parser.parse()
    
    if not rule_sets:
        error_msg = "‚ùå –ù–µ –Ω–∞–π–¥–µ–Ω–æ –Ω–∏ –æ–¥–Ω–æ–π —Å—Ç—Ä–æ–∫–∏ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏"
        print(error_msg)
        logger.error(error_msg)
        return
    
    print(f"‚úÖ –°–æ–∑–¥–∞–Ω–æ {len(rule_sets)} –Ω–∞–±–æ—Ä–æ–≤ –ø—Ä–∞–≤–∏–ª")
    
    # –í—ã–≤–æ–¥–∏–º –ø—Ä–∏–º–µ—Ä—ã
    print("\nüìã –ü—Ä–∏–º–µ—Ä—ã —Å–æ–∑–¥–∞–Ω–Ω—ã—Ö –ø—Ä–∞–≤–∏–ª:")
    for i, rule_set in enumerate(rule_sets[:3], 1):
        print(f"\n   –ê—Ä—Ç–∏–∫—É–ª: {rule_set.article} | –¶–µ–Ω–∞: {rule_set.price}")
        print(f"   ‚Ä¢ –ü—Ä–∞–≤–∏–ª–æ 0: {rule_set.rule_0}")
        print(f"   ‚Ä¢ –ü—Ä–∞–≤–∏–ª–æ 1: {rule_set.rule_1}")
        print(f"   ‚Ä¢ –ü—Ä–∞–≤–∏–ª–æ 1-1: {rule_set.rule_1_1}")
        print(f"   ‚Ä¢ –ü—Ä–∞–≤–∏–ª–æ 2: {rule_set.rule_2}")
        print(f"   ‚Ä¢ –ü—Ä–∞–≤–∏–ª–æ 2-1: {rule_set.rule_2_1}")
    
    if len(rule_sets) > 3:
        print(f"\n   ... –∏ –µ—â–µ {len(rule_sets) - 3} –∞—Ä—Ç–∏–∫—É–ª–æ–≤")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ –∞—Ä—Ç–∏–∫—É–ª–æ–≤
    articles = [rs.article for rs in rule_sets]
    
    # –ü–æ–¥–∫–ª—é—á–∞–µ–º—Å—è –∫ API
    print("\n" + "="*80)
    print("üåê –®–∞–≥ 2: –ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ –∫ API")
    print("="*80)
    
    async with DiscountRulesAPI(Config()) as api:
        # –ê–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—è
        print("\nüîê –ê–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—è...")
        if not await api.login():
            error_msg = "‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –∞–≤—Ç–æ—Ä–∏–∑–æ–≤–∞—Ç—å—Å—è –≤ —Å–∏—Å—Ç–µ–º–µ"
            print(error_msg)
            logger.error(error_msg)
            return
        print("‚úÖ –ê–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—è —É—Å–ø–µ—à–Ω–∞")
        
        # –ü–æ–ª—É—á–∞–µ–º –ø—Ä–∞–≤–∏–ª–∞
        print(f"\nüîç –ü–æ–∏—Å–∫ –ø—Ä–∞–≤–∏–ª –¥–ª—è {len(articles)} –∞—Ä—Ç–∏–∫—É–ª–æ–≤...")
        rules_by_article = await api.find_rules_by_articles(articles)
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è
        print("\n" + "="*80)
        print("‚úì –®–∞–≥ 3: –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø—Ä–∞–≤–∏–ª —á–µ—Ä–µ–∑ API")
        print("="*80)
        
        validator = RulesValidator(api, terminal_id=1541)
        
        total_articles = len(rule_sets)
        for idx, rule_set in enumerate(rule_sets, 1):
            print(f"\n[{idx}/{total_articles}] –ü—Ä–æ–≤–µ—Ä–∫–∞ –∞—Ä—Ç–∏–∫—É–ª–∞ {rule_set.article}...")
            
            api_rules = rules_by_article.get(rule_set.article, [])
            result = await validator.validate(rule_set, api_rules)
            validator.results.append(result)
            
            if result['status'] == 'NO_API_RULES':
                print(f"   ‚ö†Ô∏è  {result['message']}")
            else:
                print(f"   üìä {result['message']}")
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ Excel
    print("\n" + "="*80)
    print("üíæ –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤...")
    print("="*80)
    
    excel_file = validator.export_to_excel("validation_results.xlsx")
    print(f"‚úÖ –§–∞–π–ª —Å–æ—Ö—Ä–∞–Ω–µ–Ω: {excel_file}")
    
    # –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
    print("\n" + "="*80)
    print("üìä –ò–¢–û–ì–û–í–ê–Ø –°–¢–ê–¢–ò–°–¢–ò–ö–ê")
    print("="*80)
    
    total = len(validator.results)
    with_rules = sum(1 for r in validator.results if r['status'] == 'COMPLETED')
    without_rules = sum(1 for r in validator.results if r['status'] == 'NO_API_RULES')
    
    total_ok = sum(r.get('ok_count', 0) for r in validator.results)
    total_fail = sum(r.get('fail_count', 0) for r in validator.results)
    total_error = sum(r.get('error_count', 0) for r in validator.results)
    
    print(f"\nüì¶ –í—Å–µ–≥–æ –∞—Ä—Ç–∏–∫—É–ª–æ–≤: {total}")
    print(f"‚úÖ –ü—Ä–æ–≤–µ—Ä–µ–Ω–æ: {with_rules}")
    print(f"‚ö†Ô∏è  –ë–µ–∑ –ø—Ä–∞–≤–∏–ª –≤ API: {without_rules}")
    print(f"\nüéØ –ü—Ä–æ–≤–µ—Ä–∫–∏ –ø—Ä–∞–≤–∏–ª:")
    print(f"   ‚úÖ –£—Å–ø–µ—à–Ω–æ: {total_ok}")
    print(f"   ‚ùå –û—à–∏–±–∫–∏: {total_fail}")
    print(f"   ‚ö†Ô∏è  API –æ—à–∏–±–∫–∏: {total_error}")
    
    print("\n" + "="*80)
    print("‚úÖ –ü–†–û–ì–†–ê–ú–ú–ê –ó–ê–í–ï–†–®–ï–ù–ê")
    print("="*80)
    
    logger.info("\n" + "="*80)
    logger.info("–ò–¢–û–ì–û–í–ê–Ø –°–¢–ê–¢–ò–°–¢–ò–ö–ê")
    logger.info("="*80)
    logger.info(f"–í—Å–µ–≥–æ –∞—Ä—Ç–∏–∫—É–ª–æ–≤: {total}")
    logger.info(f"–ü—Ä–æ–≤–µ—Ä–µ–Ω–æ: {with_rules}")
    logger.info(f"–ë–µ–∑ –ø—Ä–∞–≤–∏–ª: {without_rules}")
    logger.info(f"–£—Å–ø–µ—à–Ω—ã—Ö –ø—Ä–æ–≤–µ—Ä–æ–∫: {total_ok}")
    logger.info(f"–û—à–∏–±–æ–∫: {total_fail}")
    logger.info(f"API –æ—à–∏–±–æ–∫: {total_error}")
    logger.info("="*80)
    logger.info("–ü—Ä–æ–≥—Ä–∞–º–º–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞")
    logger.info("="*80)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("\n–ü—Ä–æ–≥—Ä–∞–º–º–∞ –ø—Ä–µ—Ä–≤–∞–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
    except Exception as e:
        logger.error(f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞: {e}", exc_info=True)
        
        
# source .venv/Scripts/activate
# python discount_checker.py