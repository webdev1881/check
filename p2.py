import asyncio
import aiohttp
import ssl
import pandas as pd
import random
from typing import List, Dict, Tuple
from dataclasses import dataclass
import logging
from pathlib import Path


# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('discount_checker.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class Config:
    BASE_URL = "https://89.105.216.114"
    USERNAME = "Yulia"
    PASSWORD = "SY1804$@"
    
    BATCH_SIZE = 100
    USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36'
    
    EXCEL_FILE = "data.xlsx"  # Имя Excel файла в корне приложения


@dataclass
class RuleSet:
    """Набор правил для одной строки"""
    article: str  # Артикул из столбца C
    price: float  # Цена из столбца I
    
    # Уровни
    level_0: float  # Случайное число от 0 до K (не включая)
    level_1: float  # Случайное число от K (включая) до P (не включая)
    level_2: float  # Случайное число выше на 150-300% от P (включая)
    
    # Правила
    rule_0: float  # level_0 * price
    rule_1: float  # (price - L) * level_1
    rule_1_1: float  # Q * K
    rule_2: float  # (price - P) * level_2
    rule_2_1: float  # Q * P
    
    # Исходные данные для справки
    k_value: float
    l_value: float
    p_value: float
    q_value: float


class ExcelParser:
    """Парсер Excel файла"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        
    def parse(self) -> List[RuleSet]:
        """Парсит Excel и создает наборы правил для каждой строки"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(self.file_path)
            logger.info(f"Excel файл загружен: {len(df)} строк")
            
            rule_sets = []
            
            for index, row in df.iterrows():
                # Проверяем, что столбец C не пустой
                if pd.isna(row.iloc[2]) or str(row.iloc[2]).strip() == '':
                    continue
                
                try:
                    # Извлекаем значения из столбцов
                    article = str(row.iloc[2]).strip()  # Столбец C
                    price = float(row.iloc[8])  # Столбец I
                    k_value = float(row.iloc[10])  # Столбец K
                    l_value = float(row.iloc[11])  # Столбец L
                    p_value = float(row.iloc[15])  # Столбец P
                    q_value = float(row.iloc[16])  # Столбец Q
                    
                    # Создаем уровни (предварительные расчеты)
                    level_0 = round(random.uniform(0, k_value), 2)
                    level_1 = round(random.uniform(k_value, p_value), 2)
                    level_2 = round(p_value * random.uniform(1.5, 3.0), 2)
                    
                    # Применяем правила
                    rule_0 = round(level_0 * price, 2)
                    rule_1 = round((price - l_value) * level_1, 2)
                    rule_1_1 = round((price - l_value) * k_value, 2)
                    rule_2 = round((price - q_value) * level_2, 2)
                    rule_2_1 = round((price - q_value) * p_value, 2)
                    
                    rule_set = RuleSet(
                        article=article,
                        price=price,
                        level_0=level_0,
                        level_1=level_1,
                        level_2=level_2,
                        rule_0=rule_0,
                        rule_1=rule_1,
                        rule_1_1=rule_1_1,
                        rule_2=rule_2,
                        rule_2_1=rule_2_1,
                        k_value=k_value,
                        l_value=l_value,
                        p_value=p_value,
                        q_value=q_value
                    )
                    
                    rule_sets.append(rule_set)
                    logger.debug(f"Создан набор правил для артикула {article}")
                    
                except (ValueError, IndexError) as e:
                    logger.warning(f"Ошибка обработки строки {index + 1}: {e}")
                    continue
            
            logger.info(f"Создано {len(rule_sets)} наборов правил")
            return rule_sets
            
        except Exception as e:
            logger.error(f"Ошибка чтения Excel файла: {e}")
            raise


class DiscountRulesAPI:
    """API клиент для работы с системой скидок"""
    
    def __init__(self, config: Config):
        self.config = config
        self.session = None
        self.cookies = None
        
    async def __aenter__(self):
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        self.session = aiohttp.ClientSession(connector=connector)
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            await self.session.close()
            await asyncio.sleep(0.25)  # Даем время на закрытие
    
    async def login(self) -> bool:
        """Авторизация в системе"""
        url = f"{self.config.BASE_URL}/api/login"
        payload = {
            "username": self.config.USERNAME,
            "password": self.config.PASSWORD
        }
        
        try:
            async with self.session.post(url, json=payload) as response:
                if response.status == 200:
                    self.cookies = response.cookies
                    logger.info("Авторизация успешна")
                    return True
                else:
                    text = await response.text()
                    logger.error(f"Ошибка авторизации: {response.status} - {text}")
                    return False
        except Exception as e:
            logger.error(f"Ошибка при авторизации: {e}")
            return False
    
    async def get_discount_rules_page(self, offset: int = 0) -> Tuple[List[Dict], int]:
        """Получает одну страницу правил скидок"""
        url = f"{self.config.BASE_URL}/discountRule/list"
        
        payload = {
            "count": self.config.BATCH_SIZE,
            "filter": {},
            "offset": offset,
            "period": {},
            "sort": {
                "fields": [
                    {
                        "field": "name",
                        "asc": True
                    }
                ]
            }
        }
        
        headers = {
            'accept': '*/*',
            'content-type': 'application/json',
            'origin': self.config.BASE_URL,
            'referer': f"{self.config.BASE_URL}/",
            'user-agent': self.config.USER_AGENT
        }
        
        try:
            async with self.session.post(url, json=payload, headers=headers, cookies=self.cookies) as response:
                if response.status == 200:
                    data = await response.json()
                    return data.get('data', []), data.get('count', 0)
                else:
                    text = await response.text()
                    logger.error(f"Ошибка получения данных: {response.status} - {text}")
                    return [], 0
        except Exception as e:
            logger.error(f"Ошибка при запросе данных: {e}")
            return [], 0
    
    async def get_all_discount_rules(self) -> List[Dict]:
        """Получает все правила скидок с учетом пагинации"""
        all_rules = []
        offset = 0
        
        while True:
            rules, total_count = await self.get_discount_rules_page(offset)
            
            if not rules:
                break
            
            all_rules.extend(rules)
            logger.info(f"Загружено {len(all_rules)} из {total_count} правил")
            
            if len(all_rules) >= total_count:
                break
            
            offset += self.config.BATCH_SIZE
        
        logger.info(f"Всего загружено {len(all_rules)} правил")
        return all_rules
    
    async def find_rules_by_articles(self, articles: List[str]) -> Dict[str, List[Dict]]:
        """Находит правила для списка артикулов (два правила на артикул с приоритетами 55 и 50)"""
        all_rules = await self.get_all_discount_rules()
        
        # Группируем правила по артикулам
        rules_by_article = {article: [] for article in articles}
        
        for rule in all_rules:
            name = rule.get('name', '')
            
            # Проверяем структуру "Ахтирка_{Article}" (игнорируем суффиксы типа "_ц3")
            if name.startswith('Ахтирка_'):
                # Убираем префикс "Ахтирка_"
                name_without_prefix = name.split('Ахтирка_', 1)[1]
                
                # Берем только первое подчеркивание (игнорируем суффиксы)
                article = name_without_prefix.split('_')[0]
                
                if article in rules_by_article:
                    rules_by_article[article].append(rule)
        
        # Логируем результаты и проверяем наличие правил с приоритетами 55 и 50
        for article, rules in rules_by_article.items():
            if rules:
                priorities = [r.get('priority') for r in rules]
                logger.info(f"Для артикула {article} найдено {len(rules)} правил с приоритетами: {priorities}")
                
                # Проверяем наличие обоих приоритетов
                has_55 = any(p == 55 for p in priorities)
                has_50 = any(p == 50 for p in priorities)
                
                if not (has_55 and has_50):
                    logger.warning(f"⚠️ Артикул {article}: отсутствует правило с приоритетом {'55' if not has_55 else '50'}")
            else:
                logger.warning(f"Для артикула {article} не найдено правил")
        
        return rules_by_article
    
    async def test_discount_rule(self, article: str, quantity: float, price: float, terminal_id: int = 1541) -> Dict:
        """Тестирует правило скидки"""
        url = f"{self.config.BASE_URL}/discountRuleTester/process"
        
        payload = {
            "items": [
                {
                    "extSku": {
                        "id": article
                    },
                    "quantity": quantity,
                    "price": str(price), 
                    "discount": 0,
                    "coupons": [],
                    "paidByPoints": 0,
                    "appliedDiscountAmount": 0,
                    "isFullTank": False,
                    "amount": round(quantity * price, 2)
                }
            ],
            "promoCodes": "",
            "cardCode": None,
            "clientId": None,
            "payFormType": 0,
            "terminalId": terminal_id,
            "date": "2025-11-04T18:32:03.496Z"
        }
        
        headers = {
            'accept': '*/*',
            'content-type': 'application/json',
            'origin': self.config.BASE_URL,
            'referer': f"{self.config.BASE_URL}/",
            'user-agent': self.config.USER_AGENT
        }
        
        logger.info(f"\n📤 Запрос к discountRuleTester/process:")
        logger.info(f"   Article: {article}, Quantity: {quantity}, Price: {price}, TerminalId: {terminal_id}")
        logger.info(f"   Payload: {payload}")
        
        try:
            async with self.session.post(url, json=payload, headers=headers, cookies=self.cookies) as response:
                logger.info(f"\n📥 Ответ от API:")
                logger.info(f"   HTTP Status: {response.status}")
                
                if response.status == 200:
                    try:
                        data = await response.json()
                        # logger.info(f"   Response JSON: {data}")
                        
                        # ВАЖНО: Проверяем наличие ошибки в ответе
                        if data and isinstance(data, dict):
                            # Если есть поле error и оно не пустое - это ошибка
                            error_msg = data.get('error')
                            if error_msg and error_msg != '':
                                logger.error(f"   ❌ API вернул ошибку: {error_msg}")
                                return {
                                    'success': False,
                                    'error': error_msg,
                                    'total_discount': 0
                                }
                            
                            # Если data = None - это тоже ошибка
                            data_obj = data.get('data')
                            if data_obj is None:
                                logger.warning(f"   ⚠️  API вернул data=None (error: {error_msg})")
                                return {
                                    'success': False,
                                    'error': error_msg if error_msg else 'API вернул data=None',
                                    'total_discount': 0
                                }
                        
                    except Exception as json_error:
                        logger.error(f"   ❌ Ошибка парсинга JSON: {json_error}")
                        text = await response.text()
                        logger.error(f"   Текст ответа: {text}")
                        return {
                            'success': False,
                            'error': f'JSON parse error: {str(json_error)[:200]}',
                            'total_discount': 0
                        }
                    
                    # Безопасное извлечение totalDiscountAmount
                    total_discount = 0
                    if isinstance(data, dict):
                        data_obj = data.get('data')
                        if isinstance(data_obj, dict):
                            total_discount = data_obj.get('totalDiscountAmount', 0)
                            logger.info(f"   ✅ Успешно: totalDiscountAmount = {total_discount}")
                    
                    # Преобразуем в float если нужно
                    try:
                        total_discount = float(total_discount) if total_discount else 0
                    except (ValueError, TypeError) as e:
                        logger.error(f"   ❌ Ошибка преобразования в float: {e}")
                        total_discount = 0
                    
                    return {
                        'success': True,
                        'data': data,
                        'total_discount': total_discount
                    }
                else:
                    text = await response.text()
                    logger.error(f"   ❌ HTTP ошибка {response.status}")
                    logger.error(f"   Текст ответа: {text}")
                    
                    # Проверяем на ошибку БД - артикул не найден
                    if 'is not present in table' in text or 'ext_sku_group_id' in text:
                        return {
                            'success': False,
                            'error': 'Артикул не найден в системе',
                            'total_discount': 0
                        }
                    return {
                        'success': False,
                        'error': text[:200],  # Ограничиваем длину
                        'total_discount': 0
                    }
        except Exception as e:
            logger.error(f"   ❌ Исключение: {type(e).__name__}: {e}")
            return {
                'success': False,
                'error': f'{type(e).__name__}: {str(e)[:200]}',
                'total_discount': 0
            }


@dataclass
class ValidationCheck:
    """Результат проверки одного правила"""
    rule_name: str
    quantity: float
    price_without_discount: float
    price_with_discount: float
    expected_discount: float
    actual_discount: float
    difference: float
    status: str  # 'OK', 'FAIL', 'ERROR'
    error: str = None


class RulesValidator:
    """Валидатор правил"""
    
    def __init__(self, api: DiscountRulesAPI, terminal_id: int = 1541):
        self.api = api
        self.terminal_id = terminal_id
        self.results = []
    
    async def validate(self, rule_set: RuleSet, api_rules: List[Dict]) -> Dict:
        """Проверяет правила из API против расчетных правил"""
        validation_result = {
            'article': rule_set.article,
            'price': rule_set.price,
            'api_rules_count': len(api_rules),
            'checks': []
        }
        
        if not api_rules:
            validation_result['status'] = 'NO_API_RULES'
            validation_result['message'] = 'Правила не найдены в API'
            return validation_result
        
        logger.info(f"\n{'='*80}")
        logger.info(f"🔍 Проверка артикула: {rule_set.article}")
        logger.info(f"Найдено правил в API: {len(api_rules)}")
        logger.info(f"{'='*80}")
        
        # Проверяем каждое из 5 правил
        rules_to_check = [
            ('Правило 0', rule_set.level_0, rule_set.rule_0),
            ('Правило 1', rule_set.level_1, rule_set.rule_1),
            ('Правило 1-1', rule_set.k_value, rule_set.rule_1_1),
            ('Правило 2', rule_set.level_2, rule_set.rule_2),
            ('Правило 2-1', rule_set.p_value, rule_set.rule_2_1),
        ]
        
        for rule_name, quantity, price_with_discount in rules_to_check:
            # Цена без скидки
            price_without_discount = round(quantity * rule_set.price, 2)
            
            # Ожидаемая скидка = цена без скидки - цена со скидкой
            expected_discount = round(price_without_discount - price_with_discount, 2)
            
            logger.info(f"\n📋 {rule_name}:")
            logger.info(f"   Количество ОТ: {quantity}")
            logger.info(f"   Цена без скидки: {price_without_discount}")
            logger.info(f"   Ожидаемая цена со скидкой: {price_with_discount}")
            logger.info(f"   Ожидаемая скидка: {expected_discount}")
            
            # Ищем пару правил в API (приоритет 55 и 50)
            found_rules = self._find_matching_rules(api_rules, quantity, rule_set.price)
            
            if found_rules:
                logger.info(f"   📌 Найдены правила в API:")
                for fr in found_rules:
                    priority = fr.get('priority', 'N/A')
                    name = fr.get('name', 'N/A')
                    logger.info(f"      • Priority {priority}: {name}")
            else:
                logger.warning(f"   ⚠️  Правила с приоритетом 55/50 не найдены")
            
            # Тестируем через API
            result = await self.api.test_discount_rule(
                article=rule_set.article,
                quantity=quantity,
                price=rule_set.price,
                terminal_id=self.terminal_id
            )
            
            if result['success']:
                actual_discount = result['total_discount']
                actual_price_with_discount = round(price_without_discount - actual_discount, 2)
                difference = abs(expected_discount - actual_discount)
                
                # Определяем статус (допуск 0.01)
                if not found_rules:
                    status = 'NOT_FOUND'
                else:
                    status = 'OK' if difference <= 0.01 else 'FAIL'
                
                check = ValidationCheck(
                    rule_name=rule_name,
                    quantity=quantity,
                    price_without_discount=price_without_discount,
                    price_with_discount=price_with_discount,
                    expected_discount=expected_discount,
                    actual_discount=actual_discount,
                    difference=difference,
                    status=status
                )
                
                # Красивый вывод
                if status == 'OK':
                    logger.info(f"   ✅ API скидка: {actual_discount} (цена: {actual_price_with_discount}) - СОВПАДАЕТ")
                elif status == 'NOT_FOUND':
                    logger.warning(f"   ⚠️  API скидка: {actual_discount} (цена: {actual_price_with_discount}) - ПРАВИЛА НЕ НАЙДЕНЫ")
                else:
                    logger.warning(f"   ❌ API скидка: {actual_discount} (цена: {actual_price_with_discount}) - РАСХОЖДЕНИЕ {difference}")
                
            else:
                check = ValidationCheck(
                    rule_name=rule_name,
                    quantity=quantity,
                    price_without_discount=price_without_discount,
                    price_with_discount=price_with_discount,
                    expected_discount=expected_discount,
                    actual_discount=0,
                    difference=expected_discount,
                    status='ERROR',
                    error=result.get('error', 'Unknown error')
                )
                logger.error(f"   ❌ Ошибка API: {check.error}")
            
            validation_result['checks'].append(check)
        
        # Подсчет статистики
        ok_count = sum(1 for c in validation_result['checks'] if c.status == 'OK')
        fail_count = sum(1 for c in validation_result['checks'] if c.status == 'FAIL')
        error_count = sum(1 for c in validation_result['checks'] if c.status == 'ERROR')
        not_found_count = sum(1 for c in validation_result['checks'] if c.status == 'NOT_FOUND')
        
        validation_result['status'] = 'COMPLETED'
        validation_result['ok_count'] = ok_count
        validation_result['fail_count'] = fail_count
        validation_result['error_count'] = error_count
        validation_result['not_found_count'] = not_found_count
        validation_result['message'] = f'Проверено 5 правил: ✅ {ok_count} | ❌ {fail_count} | ⚠️ {error_count} | 🔍 {not_found_count}'
        
        logger.info(f"\n📊 Итог: {validation_result['message']}")
        
        return validation_result
    
    def _find_matching_rules(self, api_rules: List[Dict], quantity_from: float, price: float) -> List[Dict]:
        """Находит правила с приоритетом 55 (с ДО) и 50 (без ДО) для заданного количества"""
        matching_rules = []
        
        if not api_rules:
            return matching_rules
        
        for rule in api_rules:
            if not rule or not isinstance(rule, dict):
                continue
                
            priority = rule.get('priority', 0)
            
            # Проверяем только правила с приоритетом 55 или 50
            if priority not in [55, 50]:
                continue
            
            # Ищем в resultScaleItems
            result_scale_items = rule.get('resultScaleItems', [])
            if not result_scale_items:
                continue
                
            for scale_item in result_scale_items:
                if not scale_item or not isinstance(scale_item, dict):
                    continue
                    
                results = scale_item.get('results', [])
                if not results:
                    continue
                    
                for result_item in results:
                    if not result_item or not isinstance(result_item, dict):
                        continue
                        
                    restriction = result_item.get('restriction')
                    if not restriction or not isinstance(restriction, dict):
                        continue
                        
                    conditions = restriction.get('conditions', [])
                    if not conditions:
                        continue
                    
                    # Ищем условия типа 6 (не менше) и 1 (не більше)
                    has_from = False
                    has_to = False
                    from_value = None
                    to_value = None
                    
                    for condition in conditions:
                        if not condition or not isinstance(condition, dict):
                            continue
                            
                        cond_type = condition.get('type')
                        cond_value_str = condition.get('value')
                        
                        if cond_value_str is None:
                            continue
                            
                        try:
                            cond_value = float(cond_value_str)
                        except (ValueError, TypeError):
                            continue
                        
                        if cond_type == 6:  # не менше (ОТ)
                            has_from = True
                            from_value = cond_value
                        elif cond_type == 1:  # не більше (ДО)
                            has_to = True
                            to_value = cond_value
                    
                    # Приоритет 55 должен иметь оба условия (ОТ-ДО)
                    if priority == 55 and has_from and has_to:
                        if from_value is not None and to_value is not None:
                            if from_value <= quantity_from <= to_value:
                                matching_rules.append(rule)
                                break
                    
                    # Приоритет 50 должен иметь только условие ОТ
                    elif priority == 50 and has_from and not has_to:
                        if from_value is not None:
                            if from_value <= quantity_from:
                                matching_rules.append(rule)
                                break
        
        return matching_rules
    
    def export_to_excel(self, filename: str = "validation_results.xlsx"):
        """Экспортирует результаты в Excel"""
        rows = []
        
        for result in self.results:
            article = result['article']
            price = result['price']
            status = result['status']
            
            if status == 'NO_API_RULES':
                rows.append({
                    'Артикул': article,
                    'Цена': price,
                    'Статус': 'Нет правил в API',
                    'Правило': '',
                    'Количество': '',
                    'Сумма без скидки': '',
                    'Сумма со скидкой': '',
                    'Ожидаемая скидка': '',
                    'Фактическая скидка (API)': '',
                    'Расхождение': '',
                    'Результат': '',
                    'Ошибка': ''
                })
            else:
                for check in result['checks']:
                    # Восстанавливаем детали из check
                    quantity = check.quantity if hasattr(check, 'quantity') else ''
                    price_without = check.price_without_discount if hasattr(check, 'price_without_discount') else ''
                    price_with = check.price_with_discount if hasattr(check, 'price_with_discount') else ''
                    
                    rows.append({
                        'Артикул': article,
                        'Цена': price,
                        'Статус': status,
                        'Правило': check.rule_name,
                        'Количество': quantity,
                        'Сумма без скидки': price_without,
                        'Сумма со скидкой': price_with,
                        'Ожидаемая скидка': check.expected_discount,
                        'Фактическая скидка (API)': check.actual_discount,
                        'Расхождение': check.difference,
                        'Результат': check.status,
                        'Ошибка': check.error if check.error else ''
                    })
        
        df = pd.DataFrame(rows)
        
        # Форматирование
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Результаты')
            
            # Получаем worksheet для форматирования
            worksheet = writer.sheets['Результаты']
            
            # Устанавливаем ширину колонок
            worksheet.column_dimensions['A'].width = 15  # Артикул
            worksheet.column_dimensions['B'].width = 10  # Цена
            worksheet.column_dimensions['C'].width = 20  # Статус
            worksheet.column_dimensions['D'].width = 15  # Правило
            worksheet.column_dimensions['E'].width = 12  # Количество
            worksheet.column_dimensions['F'].width = 18  # Сумма без скидки
            worksheet.column_dimensions['G'].width = 18  # Сумма со скидкой
            worksheet.column_dimensions['H'].width = 18  # Ожидаемая скидка
            worksheet.column_dimensions['I'].width = 22  # Фактическая скидка
            worksheet.column_dimensions['J'].width = 15  # Расхождение
            worksheet.column_dimensions['K'].width = 12  # Результат
            worksheet.column_dimensions['L'].width = 30  # Ошибка
            
            # Форматируем заголовки
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Форматируем результаты (цветами)
            for row in range(2, len(df) + 2):
                result_cell = worksheet.cell(row=row, column=11)  # Колонка K (Результат)
                
                if result_cell.value == 'OK':
                    result_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    result_cell.font = Font(color='006100')
                elif result_cell.value == 'FAIL':
                    result_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    result_cell.font = Font(color='9C0006')
                elif result_cell.value == 'ERROR':
                    result_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    result_cell.font = Font(color='9C6500')
        
        logger.info(f"\n💾 Результаты сохранены в {filename}")
        return filename


async def main():
    """Основная функция программы"""
    print("\n" + "="*80)
    print("🚀 ПРОГРАММА ПРОВЕРКИ ПРАВИЛ СКИДОК")
    print("="*80)
    
    logger.info("="*80)
    logger.info("Запуск программы проверки правил скидок")
    logger.info("="*80)
    
    # Проверяем наличие Excel файла
    excel_path = Path(Config.EXCEL_FILE)
    if not excel_path.exists():
        error_msg = f"❌ Excel файл не найден: {excel_path.absolute()}"
        print(error_msg)
        logger.error(error_msg)
        print("\n💡 Поместите файл 'data.xlsx' в корень приложения")
        return
    
    # Парсим Excel и создаем наборы правил
    print("\n📂 Шаг 1: Парсинг Excel файла...")
    logger.info("Шаг 1: Парсинг Excel файла")
    
    parser = ExcelParser(Config.EXCEL_FILE)
    rule_sets = parser.parse()
    
    if not rule_sets:
        error_msg = "❌ Не найдено ни одной строки для обработки"
        print(error_msg)
        logger.error(error_msg)
        return
    
    print(f"✅ Создано {len(rule_sets)} наборов правил")
    
    # Выводим примеры
    print("\n📋 Примеры созданных правил:")
    for i, rule_set in enumerate(rule_sets[:3], 1):
        print(f"\n   Артикул: {rule_set.article} | Цена: {rule_set.price}")
        print(f"   • Правило 0: {rule_set.rule_0}")
        print(f"   • Правило 1: {rule_set.rule_1}")
        print(f"   • Правило 1-1: {rule_set.rule_1_1}")
        print(f"   • Правило 2: {rule_set.rule_2}")
        print(f"   • Правило 2-1: {rule_set.rule_2_1}")
    
    if len(rule_sets) > 3:
        print(f"\n   ... и еще {len(rule_sets) - 3} артикулов")
    
    # Получаем список артикулов
    articles = [rs.article for rs in rule_sets]
    
    # Подключаемся к API
    print("\n" + "="*80)
    print("🌐 Шаг 2: Подключение к API")
    print("="*80)
    
    async with DiscountRulesAPI(Config()) as api:
        # Авторизация
        print("\n🔐 Авторизация...")
        if not await api.login():
            error_msg = "❌ Не удалось авторизоваться в системе"
            print(error_msg)
            logger.error(error_msg)
            return
        print("✅ Авторизация успешна")
        
        # Получаем правила
        # print(f"\n🔍 Поиск правил для {len(articles)} артикулов...")
        rules_by_article = await api.find_rules_by_articles(articles)
        
        # Валидация
        print("\n" + "="*80)
        print("✓ Шаг 3: Проверка правил через API")
        print("="*80)
        
        validator = RulesValidator(api, terminal_id=1541)
        
        total_articles = len(rule_sets)
        for idx, rule_set in enumerate(rule_sets, 1):
            # print(f"\n[{idx}/{total_articles}] Проверка артикула {rule_set.article}...")
            
            api_rules = rules_by_article.get(rule_set.article, [])
            result = await validator.validate(rule_set, api_rules)
            validator.results.append(result)
            
            if result['status'] == 'NO_API_RULES':
                print(f"   ⚠️  {result['message']}")
            else:
                print(f"   📊 {result['message']}")
    
    # Сохраняем в Excel
    print("\n" + "="*80)
    print("💾 Сохранение результатов...")
    print("="*80)
    
    excel_file = validator.export_to_excel("validation_results.xlsx")
    print(f"✅ Файл сохранен: {excel_file}")
    
    # Итоговая статистика
    print("\n" + "="*80)
    print("📊 ИТОГОВАЯ СТАТИСТИКА")
    print("="*80)
    
    total = len(validator.results)
    with_rules = sum(1 for r in validator.results if r['status'] == 'COMPLETED')
    without_rules = sum(1 for r in validator.results if r['status'] == 'NO_API_RULES')
    
    total_ok = sum(r.get('ok_count', 0) for r in validator.results)
    total_fail = sum(r.get('fail_count', 0) for r in validator.results)
    total_error = sum(r.get('error_count', 0) for r in validator.results)
    
    print(f"\n📦 Всего артикулов: {total}")
    print(f"✅ Проверено: {with_rules}")
    print(f"⚠️  Без правил в API: {without_rules}")
    print(f"\n🎯 Проверки правил:")
    print(f"   ✅ Успешно: {total_ok}")
    print(f"   ❌ Ошибки: {total_fail}")
    print(f"   ⚠️  API ошибки: {total_error}")
    
    print("\n" + "="*80)
    print("✅ ПРОГРАММА ЗАВЕРШЕНА")
    print("="*80)
    
    logger.info("\n" + "="*80)
    logger.info("ИТОГОВАЯ СТАТИСТИКА")
    logger.info("="*80)
    logger.info(f"Всего артикулов: {total}")
    logger.info(f"Проверено: {with_rules}")
    logger.info(f"Без правил: {without_rules}")
    logger.info(f"Успешных проверок: {total_ok}")
    logger.info(f"Ошибок: {total_fail}")
    logger.info(f"API ошибок: {total_error}")
    logger.info("="*80)
    logger.info("Программа завершена")
    logger.info("="*80)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("\nПрограмма прервана пользователем")
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}", exc_info=True)